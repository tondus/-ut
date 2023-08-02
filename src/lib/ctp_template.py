import sys 
import pandas as pd
import openpyxl
from slugify import slugify
import os
import pprint as pp
from copy import copy, deepcopy
from openpyxl.utils import rows_from_range
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.styles.borders import Border, Side
from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
from distutils.dir_util import copy_tree
from pathlib import Path
import subprocess
from os import walk
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side
import math
from datetime import datetime
from openpyxl.styles import Font

col_ricefw_id = 'ricefw_id'
col_ricefw_name = 'ricefw_name'
col_document_owner = 'document_owner'
col_date = 'date'
col_tester =  'tester'
col_program_type =  'program_type'
col_middleware =  'middleware'
col_condition_no =  'condition_no'
col_condition_desc =  'condition_desc'
col_expected_result =  'expected_result'
col_fig_title1 =  'fig_title1'
col_fig_title2 =  'fig_title2'
col_fig_title3 =  'fig_title3'
col_fig_title4 =  'fig_title4'
col_fig_title5 =  'fig_title5'
col_pass_fail =  'pass_fail'
col_actual_result =  'actual_result'

text_file = ['.xml','.json','.txt','.dat']

def replace(ws,search,replace):
    i = 0
    for r in range(1,ws.max_row+1):
        for c in range(1,ws.max_column+1):
            s = ws.cell(r,c).value
            s = str(s)
            if s != None and search in s: 
                ws.cell(r,c).value = s.replace(search,replace) 
                i += 1

def copy_cell(old_cell,new_cell,new_value):
    new_cell.value = new_value
    new_cell.font = copy(old_cell.font)
    new_cell.border = copy(old_cell.border)
    new_cell.fill = copy(old_cell.fill)
    new_cell.number_format = copy(old_cell.number_format)
    new_cell.protection = copy(old_cell.protection)
    new_cell.alignment = copy(old_cell.alignment)   

def replace_cover(program,ws):
    replace(ws,"<ricefw_id>",str(program['ricefw_id']))
    replace(ws,"<ricefw_name>",str(program['ricefw_name']))
    replace(ws,"<document_owner>",str(program['document_owner']))
    replace(ws,"<date>",get_program_date(program['date']).strftime("%d.%m.%Y"))

def replace_ctp(program,ws):
    replace(ws,"<ricefw_id>",program['ricefw_id'])
    replace(ws,"<ricefw_name>",program['ricefw_name'])
    token = program['document_owner'].split(" ")
    name = token[0]
    surname = token[1][0]
    short_name = name+"."+surname

    replace(ws,"<document_owner>",short_name)

def write_ctp_table(program,ws,wb,filename,config):
    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    
    light_green = openpyxl.styles.colors.Color(rgb='0090ee90')
    green_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=light_green)
    
    light_red = openpyxl.styles.colors.Color(rgb='00ffcccb')
    red_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=light_red)

    start_line = 8
    figure_index = 1
    for condition_no in program['test_cases'].keys():
        v = program['test_cases'][condition_no]
        condition_desc = v['condition_desc']
        steps = v['steps']
        ft = Font(name='Leelawadee', size=10)
        ft_link = Font(name='Leelawadee', size=10 , underline='single', color='0000FF')
       

        ws['A'+str(start_line)].value = condition_no
        ws['A'+str(start_line)].alignment  = Alignment(horizontal='center',vertical="center")
        ws['A'+str(start_line)].border  = thin_border
        
        ws['B'+str(start_line)].value = condition_desc
        ws['B'+str(start_line)].alignment  = Alignment(horizontal='left',vertical="center")
        ws['B'+str(start_line)].font  = ft
        ws['B'+str(start_line)].border  = thin_border
        
        for step in steps:
            ws['D'+str(start_line)].value = step['expected_result']
            ws['D'+str(start_line)].alignment  = Alignment(horizontal='left',vertical="center")
            ws['D'+str(start_line)].font  = ft
            ws['D'+str(start_line)].border  = thin_border

            ws['E'+str(start_line)].value = "Pass"
            ws['E'+str(start_line)].alignment  = Alignment(horizontal='center' ,vertical="center")
            ws['E'+str(start_line)].border  = thin_border
            ws['E'+str(start_line)].font  = ft
            ws['E'+str(start_line)].fill  = green_fill
            
           
            sheetname = 'Figure '+str(figure_index)
            link = filename+"#'"+sheetname+"'!A1"
            ws['F'+str(start_line)].font  = ft_link
            ws['F'+str(start_line)].value = sheetname
            # ws['F'+str(start_line)].style  = 'Hyperlink'
            ws['F'+str(start_line)].hyperlink = link
            ws['F'+str(start_line)].border  = thin_border
            ws['F'+str(start_line)].alignment  = Alignment(horizontal='center')

            fig_titles = [step['fig_title1'],
                          step['fig_title2'],
                          step['fig_title3'],
                          step['fig_title4'],
                          step['fig_title5']]
            add_figure(program,condition_no,condition_desc,step['expected_result'],'Figure '+str(figure_index),wb,fig_titles,config)

            start_line += 1
            figure_index += 1

def add_image(image_path,sheet,start):
    img_s = openpyxl.drawing.image.Image(image_path)
    img_s.anchor = 'B'+str(start)
    sheet.add_image(img_s)
    return img_s

def add_image_to_sheet(magick_path,image,sheet,start,line_height):
    filename = os.path.basename(image)
    image_path = magick_path+filename
    img = add_image(image_path,sheet,start)
    line = math.ceil(img.height / line_height)
    return line

def get_payload(dir_path):
    payload = None
    for path in os.listdir(dir_path):
        if os.path.isfile(os.path.join(dir_path, path)):
            _filename, extension = os.path.splitext(path)
            if extension.lower() in text_file:
                t = open(os.path.join(dir_path, path), "r", encoding="utf-8")
                payload = t.read()
                t.close()
    return payload
                   
def add_figure(program,condition_no,condition_desc,expected_result,figure_title,wb,fig_titles,config):
    start = 7
    line_height = int(config['DEFAULT']['LineHight'])
    ft = Font(name='Leelawadee', size=10)
    figure_template = wb['Figure'] 
    new_sheet = wb.copy_worksheet(figure_template)
    new_sheet.title = figure_title
    replace(new_sheet,"<condiction_no>",str(int(condition_no)))
    replace(new_sheet,"<condition_desc>",condition_desc)
    replace(new_sheet,"<expected_result>",expected_result)

    payload_path = 'source/images/'+program['ricefw_id'] + '/' + figure_title +'/'
    
    image_path = 'source/image_magick/'+program['ricefw_id'] + '/' + figure_title +'/'
    
    images = []
    if os.path.isdir(image_path): 
        for path in os.listdir(image_path):
            if os.path.isfile(os.path.join(image_path, path)):
                images.append(path)
    images.sort()
    
   
    for idx ,image in enumerate(images):
        if idx < len(fig_titles):
            title = fig_titles[idx]
            if isinstance(title, str):
                token = title.split('|')
                if len(token) > 1:
                    upper = token[0]
                    lower = token[1]
                    new_sheet['B'+str(start)].value = upper.strip()
                    new_sheet['B'+str(start)].font = Font(bold=True,name='Leelawadee', size=10)
                    new_sheet['B'+str(start+1)].value = lower.strip()
                    new_sheet['B'+str(start+1)].font = Font(name='Leelawadee', size=10)
                    start += 4
                else:
                    new_sheet['B'+str(start)].value = title.strip()
                    new_sheet['B'+str(start)].font = Font(name='Leelawadee', size=10)
                    start += 3
                 
        last_line = add_image_to_sheet(image_path,image,new_sheet,start,line_height)
        start = start + last_line + 2
    
 
    payload = get_payload(payload_path) 
    if payload != None:
        token = payload.split('\n')
        for line in token:
            new_sheet['B'+str(start)].value = line
            new_sheet['B'+str(start)].font = ft
            start += 1
    
def get_program_date(program_date):
    d = program_date
    if  isinstance(program_date, str):
        d = datetime.strptime(program_date, '%d/%m/%Y')
    return d

def prepare_folder(program): 
    group = str(program['program_type']) + '_' + str(program['middleware'])
    output_path_group = 'output/'+group+'/'
    if not os.path.exists(output_path_group):
        os.mkdir(output_path_group)
    return output_path_group    

def create_document(data,config):
    template_path = 'template/template_ctp.xlsx'
    for key in data.keys():
        print(key)
        wb = openpyxl.load_workbook(template_path)
        program = data[key]
        output_path = prepare_folder(program)
        # filename = 'OREO_PIPO_CTP_'+program['ricefw_id']+" "+ program['ricefw_name']+ '_'   +  get_program_date(program['date']).strftime("%Y%m%d")+'.xlsx'
        filename = 'OREO_PIPO_CTP_'+program['ricefw_id']+" "+ program['ricefw_name']+ '_V1.0.0.xlsx'
        replace_cover(program,wb['Cover Sheet'])
        replace_ctp(program,wb['CTP'])
        write_ctp_table(program,wb['CTP'],wb,filename,config)
        
        wb.remove(wb["Figure"])
        wb.active = 1
        wb.save(output_path+filename)
        wb.close()
        
        